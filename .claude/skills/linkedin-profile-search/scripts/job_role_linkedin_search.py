#!/usr/bin/env python3
"""
Job Role-Based LinkedIn Search
Finds HR, recruiters, and senior professionals relevant to a specific job role
"""

import sys
import time
import re
from datetime import datetime
from pathlib import Path
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import the PeopleFinder class
from linkedin_people_finder import PeopleFinder


def extract_years_of_experience(description):
    """Extract years of experience from profile description"""
    if not description:
        return "Not specified"

    # Look for patterns like "5 years", "5+ years", "5-7 years"
    patterns = [
        r'(\d+)\+?\s*years?',
        r'(\d+)-(\d+)\s*years?',
        r'over\s+(\d+)\s*years?',
        r'more than\s+(\d+)\s*years?'
    ]

    for pattern in patterns:
        match = re.search(pattern, description.lower())
        if match:
            if len(match.groups()) > 1:
                # Range found (e.g., "5-7 years")
                return f"{match.group(1)}-{match.group(2)} years"
            else:
                return f"{match.group(1)}+ years"

    return "Not specified"


def determine_role_type(title, description):
    """Determine if the person is HR, Recruiter, or Senior position"""
    title_lower = title.lower()
    desc_lower = description.lower() if description else ""

    # HR keywords
    hr_keywords = ['human resources', 'hr manager', 'hr director', 'hr business partner',
                   'people operations', 'people & culture', 'talent acquisition']

    # Recruiter keywords
    recruiter_keywords = ['recruiter', 'talent acquisition', 'recruiting', 'headhunter',
                         'sourcing', 'staffing']

    # Senior/Leadership keywords
    senior_keywords = ['senior', 'lead', 'principal', 'director', 'vp', 'vice president',
                      'chief', 'head of', 'manager', 'supervisor', 'executive']

    # Check for HR
    if any(keyword in title_lower or keyword in desc_lower for keyword in hr_keywords):
        return "HR"

    # Check for Recruiter
    if any(keyword in title_lower or keyword in desc_lower for keyword in recruiter_keywords):
        return "Recruiter"

    # Check for Senior position
    if any(keyword in title_lower for keyword in senior_keywords):
        return "Senior/Leadership"

    return "Other"


def search_relevant_contacts(companies_info, job_role, limit_per_search=15):
    """Search for relevant contacts based on job role"""

    finder = PeopleFinder()
    all_results = []

    print("\n" + "="*80)
    print("  LINKEDIN JOB APPLICATION CONTACTS SEARCH")
    print("="*80 + "\n")
    print(f"Target Job Role: {job_role}")
    print(f"Searching for: HR, Recruiters, and Senior {job_role} professionals\n")

    for i, company_info in enumerate(companies_info, 1):
        company = company_info['company']
        location = company_info['location']

        print(f"[{i}/{len(companies_info)}] Searching: {company}")
        print(f"    Location: {location}\n")

        # Search 1: HR and Recruiters
        print("    [1/3] Searching for HR & Recruiters...")
        try:
            hr_results = finder.search(
                company=company,
                location=location,
                keywords="HR recruiter talent acquisition",
                limit=limit_per_search,
                method="duckduckgo"
            )

            for result in hr_results:
                result['company'] = company
                result['location'] = location
                result['search_type'] = 'HR/Recruiter'
                result['role_type'] = determine_role_type(
                    result.get('name', ''),
                    result.get('description', '')
                )
                result['years_experience'] = extract_years_of_experience(
                    result.get('description', '')
                )

            all_results.extend(hr_results)
            print(f"        Found {len(hr_results)} HR/Recruiter profiles")
            time.sleep(2)

        except Exception as e:
            print(f"        Error: {e}")

        # Search 2: Senior positions in the target job role
        print(f"    [2/3] Searching for Senior {job_role}s...")
        try:
            senior_results = finder.search(
                company=company,
                location=location,
                keywords=f"Senior {job_role} Lead Principal",
                limit=limit_per_search,
                method="duckduckgo"
            )

            for result in senior_results:
                result['company'] = company
                result['location'] = location
                result['search_type'] = f'Senior {job_role}'
                result['role_type'] = determine_role_type(
                    result.get('name', ''),
                    result.get('description', '')
                )
                result['years_experience'] = extract_years_of_experience(
                    result.get('description', '')
                )

            all_results.extend(senior_results)
            print(f"        Found {len(senior_results)} Senior {job_role} profiles")
            time.sleep(2)

        except Exception as e:
            print(f"        Error: {e}")

        # Search 3: Managers and Directors
        print(f"    [3/3] Searching for Managers & Directors...")
        try:
            manager_results = finder.search(
                company=company,
                location=location,
                keywords="Manager Director VP",
                limit=limit_per_search,
                method="duckduckgo"
            )

            for result in manager_results:
                result['company'] = company
                result['location'] = location
                result['search_type'] = 'Management'
                result['role_type'] = determine_role_type(
                    result.get('name', ''),
                    result.get('description', '')
                )
                result['years_experience'] = extract_years_of_experience(
                    result.get('description', '')
                )

            all_results.extend(manager_results)
            print(f"        Found {len(manager_results)} Manager/Director profiles")

        except Exception as e:
            print(f"        Error: {e}")

        print(f"    Total profiles for {company}: {len([r for r in all_results if r['company'] == company])}\n")

        # Delay between companies
        if i < len(companies_info):
            time.sleep(3)

    # Remove duplicates
    unique_results = []
    seen_urls = set()

    for result in all_results:
        url = result.get('profile_url', '')
        if url and url not in seen_urls:
            unique_results.append(result)
            seen_urls.add(url)

    print(f"Total unique profiles found: {len(unique_results)}")

    return unique_results


def generate_enhanced_pdf(results, companies_info, job_role, output_file):
    """Generate comprehensive PDF with job application contacts"""

    doc = SimpleDocTemplate(output_file, pagesize=letter,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Title Page
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=20,
        alignment=1
    )

    title = Paragraph(f"LinkedIn Contacts for {job_role} Positions", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    subtitle = Paragraph("HR, Recruiters & Decision Makers",
                        ParagraphStyle('Subtitle', parent=styles['Heading2'],
                                     fontSize=16, alignment=1, textColor=colors.HexColor('#555')))
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3*inch))

    # Summary
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11)
    summary_text = f"<b>Target Role:</b> {job_role}<br/>"
    summary_text += f"<b>Companies Searched:</b> {len(companies_info)}<br/>"
    summary_text += f"<b>Total Contacts:</b> {len(results)}<br/>"
    summary_text += f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"

    elements.append(Paragraph(summary_text, summary_style))
    elements.append(Spacer(1, 0.3*inch))

    # Breakdown by type
    role_types = {}
    for r in results:
        role_type = r.get('role_type', 'Other')
        role_types[role_type] = role_types.get(role_type, 0) + 1

    breakdown_text = "<b>Contacts Breakdown:</b><br/>"
    for role_type, count in sorted(role_types.items(), key=lambda x: x[1], reverse=True):
        breakdown_text += f"&nbsp;&nbsp;• {role_type}: {count} contacts<br/>"

    elements.append(Paragraph(breakdown_text, summary_style))
    elements.append(PageBreak())

    # Results by Company
    subtitle_style = ParagraphStyle('Subtitle2', parent=styles['Heading2'],
                                   fontSize=14, textColor=colors.HexColor('#366092'))

    for company_info in companies_info:
        company = company_info['company']

        company_header = Paragraph(f"<b>{company}</b>", subtitle_style)
        elements.append(company_header)
        elements.append(Spacer(1, 0.1*inch))

        company_results = [r for r in results if r['company'] == company]

        if not company_results:
            elements.append(Paragraph("No contacts found", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            continue

        # Info
        info = Paragraph(
            f"<b>Location:</b> {company_info['location']}<br/>"
            f"<b>Contacts Found:</b> {len(company_results)}",
            styles['Normal']
        )
        elements.append(info)
        elements.append(Spacer(1, 0.15*inch))

        # Table
        table_data = [['Name', 'Role Type', 'Experience', 'Profile URL']]

        for profile in company_results:
            name_parts = profile['name'].split(' - ')
            name = name_parts[0][:28]

            table_data.append([
                name,
                profile.get('role_type', 'Other')[:18],
                profile.get('years_experience', 'N/A')[:15],
                profile['profile_url'][:42] + '...'
            ])

        table = Table(table_data, colWidths=[1.8*inch, 1.4*inch, 1.2*inch, 2.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

    doc.build(elements)
    print(f"[+] PDF report generated: {output_file}")


def generate_enhanced_excel(results, companies_info, job_role, output_file):
    """Generate comprehensive Excel with all contact details"""

    excel_data = []

    for result in results:
        name_parts = result['name'].split(' - ')
        name = name_parts[0]
        job_title = name_parts[1] if len(name_parts) > 1 else result.get('job_title', 'Not specified')

        excel_data.append({
            'Name': name,
            'Role in Company': job_title,
            'Role Type': result.get('role_type', 'Other'),
            'Years of Experience': result.get('years_experience', 'Not specified'),
            'Company': result['company'],
            'Location': result['location'],
            'LinkedIn Profile': result['profile_url'],
            'Search Category': result.get('search_type', 'General'),
            'Description': result.get('description', '')[:300]
        })

    df = pd.DataFrame(excel_data)

    # Sort by Company, then Role Type
    df = df.sort_values(['Company', 'Role Type'], ascending=[True, True])

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # All contacts sheet
        df.to_excel(writer, sheet_name='All Contacts', index=False)

        # Sheet by role type
        for role_type in df['Role Type'].unique():
            role_df = df[df['Role Type'] == role_type]
            # Remove invalid Excel sheet name characters
            sheet_name = role_type.replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '').replace(':', '').replace('?', '')[:31]
            role_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Sheet by company
        for company_info in companies_info:
            company = company_info['company']
            company_df = df[df['Company'] == company]
            if not company_df.empty:
                sheet_name = company[:31]
                company_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Format Excel
    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions['A'].width = 25  # Name
        ws.column_dimensions['B'].width = 35  # Role in Company
        ws.column_dimensions['C'].width = 18  # Role Type
        ws.column_dimensions['D'].width = 20  # Years of Experience
        ws.column_dimensions['E'].width = 30  # Company
        ws.column_dimensions['F'].width = 30  # Location
        ws.column_dimensions['G'].width = 50  # LinkedIn Profile
        ws.column_dimensions['H'].width = 20  # Search Category
        if 'I' in [cell.column_letter for cell in ws[1]]:
            ws.column_dimensions['I'].width = 50  # Description

    wb.save(output_file)
    print(f"[+] Excel spreadsheet generated: {output_file}")


def main():
    # Configuration
    job_role = input("\nEnter the job role you're applying for (e.g., 'Data Analyst'): ").strip()

    if not job_role:
        print("[!] Job role is required.")
        return

    # Define companies to search
    print("\n" + "="*80)
    print("Enter company details (press Enter with empty company name to finish)")
    print("="*80 + "\n")

    companies_info = []
    while True:
        company = input(f"Company {len(companies_info) + 1} name (or press Enter to finish): ").strip()
        if not company:
            break

        location = input(f"Location for {company}: ").strip()
        if not location:
            location = "United States"

        companies_info.append({
            'company': company,
            'location': location,
            'title': job_role
        })

    if not companies_info:
        print("\n[!] No companies provided. Using default companies...")
        companies_info = [
            {'company': 'Autoroboto', 'location': 'Mountain View, CA', 'title': job_role},
            {'company': 'Louisiana Primary Care Association', 'location': 'Baton Rouge, LA', 'title': job_role},
            {'company': 'Schuber Mitchell Homes', 'location': 'Joplin, MO', 'title': job_role}
        ]

    # Search
    results = search_relevant_contacts(companies_info, job_role, limit_per_search=15)

    if not results:
        print("\n[!] No contacts found.")
        return

    # Generate reports
    output_dir = Path("C:/Users/saite/OneDrive - University of Central Missouri/Documents/MS in CS/Data Analyst/Jobs")
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_file = output_dir / f"linkedin_contacts_{job_role.replace(' ', '_')}_{timestamp}.pdf"
    excel_file = output_dir / f"linkedin_contacts_{job_role.replace(' ', '_')}_{timestamp}.xlsx"

    print("\n" + "="*80)
    print("  GENERATING REPORTS")
    print("="*80 + "\n")

    generate_enhanced_pdf(results, companies_info, job_role, str(pdf_file))
    generate_enhanced_excel(results, companies_info, job_role, str(excel_file))

    # Summary
    print("\n" + "="*80)
    print("  FINAL SUMMARY")
    print("="*80)
    print(f"\nTarget Role: {job_role}")
    print(f"Total Contacts: {len(results)}")

    role_types = {}
    for r in results:
        role_type = r.get('role_type', 'Other')
        role_types[role_type] = role_types.get(role_type, 0) + 1

    print("\nBreakdown by Type:")
    for role_type, count in sorted(role_types.items(), key=lambda x: x[1], reverse=True):
        print(f"  - {role_type}: {count} contacts")

    print("\nBy Company:")
    for info in companies_info:
        company_results = [r for r in results if r['company'] == info['company']]
        print(f"  - {info['company']}: {len(company_results)} contacts")

    print("\n" + "="*80)
    print(f"\n[+] Files saved:")
    print(f"   [*] {pdf_file}")
    print(f"   [*] {excel_file}\n")


if __name__ == '__main__':
    main()
