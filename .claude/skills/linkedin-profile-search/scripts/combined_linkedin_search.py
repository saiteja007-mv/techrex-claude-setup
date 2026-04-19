#!/usr/bin/env python3
"""
Combined LinkedIn Profile Search
Searches multiple companies and combines results into a single PDF and Excel file
Uses the linkedin_people_finder.py module for actual LinkedIn scraping
"""

import sys
import time
from datetime import datetime
from pathlib import Path
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import the PeopleFinder class
from linkedin_people_finder import PeopleFinder


def search_all_companies(companies_info, keywords="Data Analyst", limit=20):
    """Search LinkedIn for all specified companies"""

    finder = PeopleFinder()
    all_results = []

    print("\n" + "="*80)
    print("  LINKEDIN PROFILE SEARCH - COMBINED REPORT")
    print("="*80 + "\n")

    for i, company_info in enumerate(companies_info, 1):
        company = company_info['company']
        location = company_info['location']

        print(f"[{i}/{len(companies_info)}] Searching: {company}")
        print(f"    Location: {location}")
        print(f"    Keywords: {keywords}")
        print("    Please wait...\n")

        try:
            # Search using DuckDuckGo
            results = finder.search(
                company=company,
                location=location,
                keywords=keywords,
                limit=limit,
                method="duckduckgo"
            )

            # Add company info to each result
            for result in results:
                result['company'] = company
                result['location'] = location
                result['title'] = keywords

                # Extract name and title from the name field
                name_full = result.get('name', 'Unknown')
                if ' - ' in name_full:
                    parts = name_full.split(' - ', 1)
                    result['name'] = parts[0].strip()
                    result['job_title'] = parts[1].strip()
                else:
                    result['name'] = name_full
                    result['job_title'] = keywords

            all_results.extend(results)

            print(f"    [+] Found {len(results)} profiles for {company}")
            print()

            # Add delay to avoid rate limiting
            if i < len(companies_info):
                time.sleep(3)

        except Exception as e:
            print(f"    [!] Error searching {company}: {e}")
            print()
            continue

    return all_results


def generate_combined_pdf(results, companies_info, output_file):
    """Generate a single PDF with all company results"""

    doc = SimpleDocTemplate(output_file, pagesize=letter,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Title Page
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=26,
        textColor=colors.HexColor('#366092'),
        spaceAfter=20,
        alignment=1  # Center
    )

    title = Paragraph("LinkedIn Profile Search Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    # Summary section
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=15
    )

    summary_title = Paragraph("Search Summary", subtitle_style)
    elements.append(summary_title)

    # Create summary info
    summary_text = f"<b>Total Companies Searched:</b> {len(companies_info)}<br/>"
    summary_text += f"<b>Total Profiles Found:</b> {len(results)}<br/>"
    summary_text += f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/><br/>"

    summary_para = Paragraph(summary_text, styles['Normal'])
    elements.append(summary_para)
    elements.append(Spacer(1, 0.3*inch))

    # Companies searched
    companies_title = Paragraph("Companies Searched:", subtitle_style)
    elements.append(companies_title)

    for i, company_info in enumerate(companies_info, 1):
        company_text = f"{i}. <b>{company_info['company']}</b> - {company_info['title']} - {company_info['location']}<br/>"
        company_para = Paragraph(company_text, styles['Normal'])
        elements.append(company_para)

    elements.append(PageBreak())

    # Results by Company
    for company_info in companies_info:
        company = company_info['company']

        # Company header
        company_header = Paragraph(f"<b>{company}</b>", subtitle_style)
        elements.append(company_header)

        # Filter results for this company
        company_results = [r for r in results if r['company'] == company]

        if not company_results:
            no_results = Paragraph(f"No profiles found for {company}", styles['Normal'])
            elements.append(no_results)
            elements.append(Spacer(1, 0.2*inch))
            continue

        # Company info
        info_text = f"<b>Position:</b> {company_info['title']}<br/>"
        info_text += f"<b>Location:</b> {company_info['location']}<br/>"
        info_text += f"<b>Profiles Found:</b> {len(company_results)}<br/>"

        info_para = Paragraph(info_text, styles['Normal'])
        elements.append(info_para)
        elements.append(Spacer(1, 0.15*inch))

        # Create table for this company's results
        table_data = [['#', 'Name', 'Title', 'Profile URL']]

        for idx, profile in enumerate(company_results, 1):
            # Truncate URL for display
            url_display = profile['profile_url'][:45] + '...' if len(profile['profile_url']) > 45 else profile['profile_url']

            table_data.append([
                str(idx),
                profile['name'][:25],
                profile.get('job_title', 'Data Analyst')[:28],
                url_display
            ])

        # Create and style table
        table = Table(table_data, colWidths=[0.4*inch, 1.8*inch, 1.8*inch, 2.5*inch])
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            # Body style
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

    # Build PDF
    doc.build(elements)
    print(f"[+] Combined PDF generated: {output_file}")


def generate_combined_excel(results, companies_info, output_file):
    """Generate Excel file with all results"""

    # Create DataFrame
    excel_data = []
    for result in results:
        excel_data.append({
            'Name': result['name'],
            'Title': result.get('job_title', 'Data Analyst'),
            'Company': result['company'],
            'Location': result['location'],
            'Profile URL': result['profile_url'],
            'Description': result.get('description', '')[:200]
        })

    df = pd.DataFrame(excel_data)

    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All Profiles', index=False)

        # Create separate sheets for each company
        for company_info in companies_info:
            company = company_info['company']
            company_df = df[df['Company'] == company]
            if not company_df.empty:
                sheet_name = company[:31]  # Excel sheet name limit
                company_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Format the Excel file
    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions['A'].width = 25  # Name
        ws.column_dimensions['B'].width = 30  # Title
        ws.column_dimensions['C'].width = 30  # Company
        ws.column_dimensions['D'].width = 35  # Location
        ws.column_dimensions['E'].width = 50  # Profile URL
        if 'F' in [cell.column_letter for cell in ws[1]]:
            ws.column_dimensions['F'].width = 40  # Description

    wb.save(output_file)
    print(f"[+] Excel file generated: {output_file}")


def main():
    # Define companies to search
    companies_info = [
        {
            'company': 'Autoroboto',
            'title': 'Data Analyst',
            'location': 'Mountain View, CA'
        },
        {
            'company': 'Louisiana Primary Care Association',
            'title': 'Data Analyst',
            'location': 'Baton Rouge, LA'
        },
        {
            'company': 'Schuber Mitchell Homes',
            'title': 'Data Analyst',
            'location': 'Joplin, MO'
        }
    ]

    # Search all companies
    results = search_all_companies(companies_info, keywords="Data Analyst", limit=20)

    if not results:
        print("\n[!] No profiles found across all companies.")
        print("[!] Try:")
        print("    - Checking company name spellings")
        print("    - Using broader location terms")
        print("    - Verifying internet connection")
        return

    # Generate outputs
    output_dir = Path("C:/Users/saite/OneDrive - University of Central Missouri/Documents/MS in CS/Data Analyst/Jobs")
    output_dir.mkdir(parents=True, exist_ok=True)

    pdf_file = output_dir / "linkedin_combined_results.pdf"
    excel_file = output_dir / "linkedin_combined_results.xlsx"

    print("\n" + "="*80)
    print("  GENERATING REPORTS")
    print("="*80 + "\n")

    # Generate PDF
    generate_combined_pdf(results, companies_info, str(pdf_file))

    # Generate Excel
    generate_combined_excel(results, companies_info, str(excel_file))

    # Summary
    print("\n" + "="*80)
    print("  RESULTS SUMMARY")
    print("="*80)

    for info in companies_info:
        company_results = [r for r in results if r['company'] == info['company']]
        print(f"\n  {info['company']}")
        print(f"  - Position: {info['title']}")
        print(f"  - Location: {info['location']}")
        print(f"  - Profiles Found: {len(company_results)}")

    print("\n" + "="*80)
    print(f"\n[+] Report generation complete!")
    print(f"\n   Files saved:")
    print(f"   [*] {pdf_file}")
    print(f"   [*] {excel_file}\n")


if __name__ == '__main__':
    main()
