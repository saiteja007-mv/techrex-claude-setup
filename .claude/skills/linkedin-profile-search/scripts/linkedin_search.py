#!/usr/bin/env python3
"""
LinkedIn Profile Search Script
Searches for LinkedIn profiles by company and location without authentication.
Exports results to terminal, Excel, and PDF formats.
"""

import argparse
import time
import random
import sys
from datetime import datetime
from urllib.parse import quote_plus
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


class LinkedInSearcher:
    """LinkedIn profile searcher using public data"""

    def __init__(self, company, location, limit=50):
        self.company = company
        self.location = location
        self.limit = limit
        self.results = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }

    def build_search_url(self, page=1):
        """Build LinkedIn search URL with company and location filters"""
        base_url = "https://www.linkedin.com/search/results/people/"

        # URL encode parameters
        company_encoded = quote_plus(self.company)
        location_encoded = quote_plus(self.location)

        # Build search URL (using Google as proxy to find public LinkedIn profiles)
        google_search = f'site:linkedin.com/in/ "{self.company}" "{self.location}"'
        google_url = f"https://www.google.com/search?q={quote_plus(google_search)}&start={page*10}"

        return google_url

    def search_profiles(self):
        """Search for LinkedIn profiles using Google search as proxy"""
        print(f"\n[*] Searching LinkedIn profiles...")
        print(f"   Company: {self.company}")
        print(f"   Location: {self.location}")
        print(f"   Limit: {self.limit}\n")

        # Use Google to find public LinkedIn profiles
        try:
            for page in range(0, (self.limit // 10) + 1):
                url = self.build_search_url(page)

                print(f"   Fetching page {page + 1}...")
                response = requests.get(url, headers=self.headers, timeout=10)

                if response.status_code == 200:
                    self.parse_google_results(response.text)
                else:
                    print(f"   ⚠️  Page {page + 1} returned status {response.status_code}")

                # Random delay to avoid rate limiting
                time.sleep(random.uniform(2, 5))

                if len(self.results) >= self.limit:
                    break

            print(f"\n[+] Found {len(self.results)} profiles\n")

        except Exception as e:
            print(f"[!] Search error: {e}")
            sys.exit(1)

    def parse_google_results(self, html):
        """Parse Google search results to extract LinkedIn profile URLs"""
        soup = BeautifulSoup(html, 'html.parser')

        # Find all result links
        for result in soup.find_all('div', class_='g'):
            if len(self.results) >= self.limit:
                break

            # Extract link
            link_tag = result.find('a')
            if link_tag and 'href' in link_tag.attrs:
                url = link_tag['href']

                # Filter for LinkedIn profile URLs
                if 'linkedin.com/in/' in url and url not in [r['profile_url'] for r in self.results]:
                    # Extract name from URL or title
                    title_tag = result.find('h3')
                    title = title_tag.get_text() if title_tag else "Unknown"

                    # Extract snippet for additional info
                    snippet_tag = result.find('div', class_='VwiC3b')
                    snippet = snippet_tag.get_text() if snippet_tag else ""

                    profile = self.extract_profile_from_snippet(url, title, snippet)
                    self.results.append(profile)

    def extract_profile_from_snippet(self, url, title, snippet):
        """Extract basic profile info from search result snippet"""
        # Clean URL
        if url.startswith('/url?q='):
            url = url.split('/url?q=')[1].split('&')[0]

        # Extract name from title (remove " - LinkedIn" suffix)
        name = title.replace(' - LinkedIn', '').replace(' | LinkedIn', '').strip()

        # Try to extract title/position from snippet
        job_title = "Not specified"
        if ' at ' in snippet:
            parts = snippet.split(' at ')
            if len(parts) > 0:
                job_title = parts[0].strip()
        elif '·' in snippet:
            parts = snippet.split('·')
            if len(parts) > 1:
                job_title = parts[0].strip()

        return {
            'name': name,
            'title': job_title,
            'company': self.company,
            'location': self.location,
            'profile_url': url
        }

    def display_terminal(self):
        """Display results in terminal as formatted table"""
        if not self.results:
            print("[!] No results to display")
            return

        print("\n" + "="*80)
        print(f"  LINKEDIN PROFILE SEARCH RESULTS")
        print("="*80)
        print(f"  Company: {self.company}")
        print(f"  Location: {self.location}")
        print(f"  Total Profiles: {len(self.results)}")
        print("="*80 + "\n")

        # Create DataFrame for pretty printing
        df = pd.DataFrame(self.results)
        df_display = df[['name', 'title', 'location']].head(20)

        # Print table
        print(df_display.to_string(index=False))

        if len(self.results) > 20:
            print(f"\n... and {len(self.results) - 20} more profiles (see Excel/PDF for full list)\n")

        print("\n" + "="*80 + "\n")

    def export_excel(self, filename):
        """Export results to Excel with formatting"""
        if not self.results:
            print("❌ No results to export")
            return

        excel_file = f"{filename}.xlsx"

        try:
            # Create DataFrame
            df = pd.DataFrame(self.results)
            df = df[['name', 'title', 'company', 'location', 'profile_url']]
            df.columns = ['Name', 'Title', 'Company', 'Location', 'Profile URL']

            # Write to Excel
            df.to_excel(excel_file, index=False, sheet_name='LinkedIn Profiles')

            # Apply formatting
            wb = load_workbook(excel_file)
            ws = wb.active

            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Column widths
            ws.column_dimensions['A'].width = 25  # Name
            ws.column_dimensions['B'].width = 35  # Title
            ws.column_dimensions['C'].width = 25  # Company
            ws.column_dimensions['D'].width = 20  # Location
            ws.column_dimensions['E'].width = 50  # Profile URL

            # Save
            wb.save(excel_file)

            print(f"[+] Excel exported: {excel_file}")

        except Exception as e:
            print(f"[!] Excel export error: {e}")

    def export_pdf(self, filename):
        """Export results to PDF report"""
        if not self.results:
            print("❌ No results to export")
            return

        pdf_file = f"{filename}.pdf"

        try:
            doc = SimpleDocTemplate(pdf_file, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1  # Center
            )
            title = Paragraph("LinkedIn Profile Search Results", title_style)
            elements.append(title)

            # Search criteria
            criteria_style = styles['Normal']
            criteria = Paragraph(
                f"<b>Company:</b> {self.company}<br/>"
                f"<b>Location:</b> {self.location}<br/>"
                f"<b>Total Profiles:</b> {len(self.results)}<br/>"
                f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                criteria_style
            )
            elements.append(criteria)
            elements.append(Spacer(1, 0.3*inch))

            # Create table data
            table_data = [['Name', 'Title', 'Location']]
            for profile in self.results:
                table_data.append([
                    profile['name'][:30],  # Truncate long names
                    profile['title'][:40],  # Truncate long titles
                    profile['location'][:25]  # Truncate long locations
                ])

            # Create table
            table = Table(table_data, colWidths=[2.0*inch, 2.8*inch, 1.7*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))

            elements.append(table)

            # Build PDF
            doc.build(elements)

            print(f"[+] PDF exported: {pdf_file}")

        except Exception as e:
            print(f"[!] PDF export error: {e}")


def main():
    parser = argparse.ArgumentParser(
        description='Search LinkedIn profiles by company and location'
    )
    parser.add_argument('--company', required=True, help='Company name to search')
    parser.add_argument('--location', required=True, help='Location to filter by')
    parser.add_argument('--output', default='linkedin_results', help='Output filename prefix')
    parser.add_argument('--limit', type=int, default=50, help='Maximum profiles to extract')

    args = parser.parse_args()

    # Create searcher
    searcher = LinkedInSearcher(args.company, args.location, args.limit)

    # Execute search
    searcher.search_profiles()

    # Display results
    searcher.display_terminal()

    # Export to Excel and PDF
    searcher.export_excel(args.output)
    searcher.export_pdf(args.output)

    print(f"\n[+] Search complete! Files saved:")
    print(f"   [*] {args.output}.xlsx")
    print(f"   [*] {args.output}.pdf\n")


if __name__ == '__main__':
    main()
